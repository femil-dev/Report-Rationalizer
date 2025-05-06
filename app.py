from flask import Flask, render_template, request, jsonify, send_from_directory
import os
from werkzeug.utils import secure_filename
import csv
import json
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

app = Flask(__name__)

# Define file upload folder and allowed extensions
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xml', 'json','xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure the upload folder exists
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Check if the file has allowed extension
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_columns_from_csv(file_path):
    with open(file_path, newline='') as csvfile:
        reader = csv.reader(csvfile)
        columns = next(reader)  # Read the header row to get the column names
        return columns

def get_data_from_xml(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    namespaces = {
        'saw': 'com.siebel.analytics.web/report/v1.1',
        'xsi': 'http://www.w3.org/2001/XMLSchema-instance',
        'sawx': 'com.siebel.analytics.web/expression/v1.1'
    }
    
    # Get the subjectArea from the saw:criteria
    dataSource= [root.find(".//saw:criteria", namespaces=namespaces).get('subjectArea').replace('"', '').lower()]
    
    # Get columns (expr elements for column)
    expr_elements_column = root.findall(".//saw:column//sawx:expr[@xsi:type='sawx:sqlExpression']", namespaces=namespaces)
    dataItem = [elem.text.replace('"', '').lower() for elem in expr_elements_column]

    # Get filters (expr elements for filter)
    expr_elements_filter = root.findall(".//saw:filter//sawx:expr[@xsi:type='sawx:sqlExpression']", namespaces=namespaces)
    dataFilter = [elem.text.replace('"', '').lower() for elem in expr_elements_filter]

    return dataItem, dataFilter, dataSource

def get_columns_from_json(file_path):
    with open(file_path) as jsonfile:
        data = json.load(jsonfile)
        columns = list(data[0].keys()) if data else []
        return columns

def get_data_from_file(file_path):
    if file_path.lower().endswith('.csv'):
        return get_columns_from_csv(file_path)
    elif file_path.lower().endswith('.xml'):
        return get_data_from_xml(file_path)
    elif file_path.lower().endswith('.json'):
        return get_columns_from_json(file_path)
    else:
        raise ValueError("Unsupported file format")

def calculate_matching_percentage(columns_file1, columns_file2):
    matching_columns = set(columns_file1).intersection(columns_file2)
    total_columns = len(set(columns_file1).union(columns_file2))
    
    if total_columns == 0:
        return 0
    
    matching_percentage = (len(matching_columns) / total_columns) * 100
    return matching_percentage

def append_to_excel(data_row, output_file, sheet_name):
    # Load the workbook or create a new one if it doesn't exist
    if os.path.exists(output_file):
        workbook = openpyxl.load_workbook(output_file)
    else:
        workbook = openpyxl.Workbook()
    
    # Remove the default sheet (if it exists) before proceeding
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Check if the sheet already exists
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # If the sheet doesn't exist, create it
        sheet = workbook.create_sheet(sheet_name)
        
        # Write main headers and subheaders for Sheet 1
        if sheet_name == "Sheet 1":
            main_headers = ["REPORT NAME", "REPORT PATH", "REPORT Datasource", "REPORT DataItem", "REPORT Data Filters"]
            subheaders = ["Name of the report", "Report Path", "Report DataSource", "Report DataItem", "Report Data Filters"]
            sheet.append(main_headers)  # Append main headers
            sheet.append(subheaders)    # Append subheaders
        
        # Write main headers and subheaders for Sheet 2
        elif sheet_name == "Sheet 2":
            main_headers = ["REPORT NAME", "Datasource Matched", "Datasource Match %", "DataItem Matched", "DataItem Match %", "Data Filters Matched", 
                            "Data Filters Match %", "Overall Match (%)"]
            subheaders = [
                "REPORTS COMBINED", 
                "Datasource Matched", 
                "Shows the percentage match between report Datasource", 
                "DataItem Matched", 
                "Shows the percentage match between report data items", 
                "Data Filters Matched", 
                "Shows the percentage match between report data filters", 
                "Shows the overall report matching percentage"
            ]
            sheet.append(main_headers)  # Append main headers
            sheet.append(subheaders)    # Append subheaders

        # Apply styles for main headers (bold, background color, and borders)
        main_header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
        main_header_font = Font(bold=True, color="000000")  # Bold and black text
        border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        # Apply to main headers (row 1)
        for cell in sheet[1]:  # The first row (main headers)
            cell.fill = main_header_fill
            cell.font = main_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align text
            cell.border = border  # Apply border to header cells

        # Apply styles for subheaders (italic, background color, and borders)
        subheader_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray background
        subheader_font = Font(italic=True, color="000000")  # Italic and black text
        for cell in sheet[2]:  # The second row (subheaders)
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Center align text
            cell.border = border  # Apply border to subheader cells

    # Write the data row to the sheet (append the data row)
    row_idx = sheet.max_row + 1  # Get the next empty row index
    for col_idx, value in enumerate(data_row, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)

        # Apply wrap text and center alignment for cells with line breaks
        if isinstance(value, str) and "\n" in value:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            # Apply center alignment for all cells (even without line breaks)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply border to all rows from the 3rd row onward
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(
                top=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

    # Autofit column widths, but only consider the longest value among those with \n
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)
        
        for cell in col:
            try:
                cell_value = str(cell.value)
                if '\n' in cell_value:
                    # Split by '\n' and get the maximum length of any individual line
                    max_length = max(max_length, max(len(line) for line in cell_value.split('\n')))
                else:
                    # Regular case for strings without '\n'
                    max_length = max(max_length, len(cell_value))
            except:
                pass
        
        # Adjust the column width by adding a little padding
        adjusted_width = max_length + 5  # Adding extra padding for header's appearance
        sheet.column_dimensions[column].width = adjusted_width

    # Increase row height for top row (main header) and subsequent rows
    sheet.row_dimensions[1].height = 40  # Increase height for header row (for more vertical space)
    sheet.row_dimensions[2].height = 35  # Increase height for subheader row

    # Adjust row heights dynamically based on the longest content in the row
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        max_row_height = 0
        for cell in row:
            if isinstance(cell.value, str):
                # Count number of lines in the cell (based on '\n' breaks)
                lines = cell.value.split('\n')
                max_row_height = max(max_row_height, len(lines))

        # Set the row height based on the maximum number of lines in any cell of that row
        if max_row_height > 0:
            sheet.row_dimensions[row[0].row].height = max_row_height * 15  # Multiply by 15 to give enough space

    workbook.save(output_file)

def compare_multiple_files(file_paths):
    # Initialize variables for matching percentages and matched data
    all_columns_data = []
    for file_path in file_paths:
        dataItem, dataFilter, dataSource = get_data_from_file(file_path)
        all_columns_data.append((dataItem, dataFilter, dataSource))

    # Find common elements across all files
    matched_dataItems_all = set(all_columns_data[0][0])
    matched_dataFilters_all = set(all_columns_data[0][1])
    matched_dataSources_all = set(all_columns_data[0][2])

    for dataItem, dataFilter, dataSource in all_columns_data[1:]:
        matched_dataItems_all.intersection_update(dataItem)
        matched_dataFilters_all.intersection_update(dataFilter)
        matched_dataSources_all.intersection_update(dataSource)

    # Collect all unique data items, filters, and sources from all files
    all_dataItems = set()
    all_dataFilters = set()
    all_dataSources = set()
    
    for dataItem, dataFilter, dataSource in all_columns_data:
        all_dataItems.update(dataItem)
        all_dataFilters.update(dataFilter)
        all_dataSources.update(dataSource)

    # Calculate the matching percentages correctly
    matching_percentage_dataItems = (len(matched_dataItems_all) / len(all_dataItems) * 100) if all_dataItems else 0
    matching_percentage_dataFilters = (len(matched_dataFilters_all) / len(all_dataFilters) * 100) if all_dataFilters else 0
    matching_percentage_dataSources = (len(matched_dataSources_all) / len(all_dataSources) * 100) if all_dataSources else 0

    overall_matching_percentage = round((matching_percentage_dataItems + matching_percentage_dataFilters + matching_percentage_dataSources) / 3, 2)

    # Rest of the function remains the same...
    # Prepare the report name and data row for Sheet 2
    file_names = [os.path.splitext(os.path.basename(file_path))[0] for file_path in file_paths]
    report_name = "_".join(file_names) + "_COMBINED"

    # Prepare the data rows for the Excel file
    data_row_sheet1 = []
    for i, file_path in enumerate(file_paths):
        file_name = file_names[i]
        dataItem, dataFilter, dataSource = all_columns_data[i]
        data_row_sheet1.append([
            file_name,
            file_path,
            ",\n".join(dataSource),
            ",\n".join(dataItem),
            ",\n".join(dataFilter)
        ])

    matched_dataItems_str = ",\n".join(matched_dataItems_all) if matched_dataItems_all else "No Match"
    matched_dataFilters_str = ",\n".join(matched_dataFilters_all) if matched_dataFilters_all else "No Match"
    matched_dataSources_str = ",\n".join(matched_dataSources_all) if matched_dataSources_all else "No Match"

    data_row_sheet2 = [
        report_name,
        matched_dataSources_str,
        matching_percentage_dataSources,
        matched_dataItems_str,
        matching_percentage_dataItems,
        matched_dataFilters_str,
        matching_percentage_dataFilters,
        overall_matching_percentage
    ]

    # Define the output file path
    output_file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{report_name}_merged.xlsx")

    # Remove existing file (this will ensure it gets overwritten)
    if os.path.exists(output_file_path):
        os.remove(output_file_path)

    # Append data to Sheet 1 and Sheet 2
    for row in data_row_sheet1:
        append_to_excel(row, output_file_path, sheet_name="Sheet 1")
    append_to_excel(data_row_sheet2, output_file_path, sheet_name="Sheet 2")

    return output_file_path, overall_matching_percentage

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files' not in request.files:
        return jsonify({'error': 'No files part'}), 400

    files = request.files.getlist('files')

    # Ensure at least two files are uploaded
    if len(files) < 2:
        return jsonify({'error': 'At least two files are required for comparison'}), 400

    # Validate that all files have allowed extensions
    for file in files:
        if file.filename == '' or not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file format or empty filename'}), 400

    # Save the uploaded files
    file_paths = []
    upload_folder = app.config['UPLOAD_FOLDER']

    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)

    for file in files:
        filename = secure_filename(file.filename)
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)
        file_paths.append(file_path)

    # Compare and generate reports for all files
    output_file_path, matching_percentage = compare_multiple_files(file_paths)

    # If the overall matching percentage is less than 75%, raise an error message
    if matching_percentage < 75:
        return render_template(
            'index.html',
            result=f"Matching percentage is {matching_percentage}. Comparison failed.",
            file_names=[os.path.basename(file.filename) for file in files],
            message_class="error"
        )

    # If the matching percentage is above or equal to 75%, show the success message and download link
    if output_file_path:
        output_file_url = f"/download/{os.path.basename(output_file_path)}"

        return render_template(
            'index.html',
            result=f"Comparison complete. Matching percentage: {matching_percentage:.2f}%",
            file_names=[os.path.basename(file.filename) for file in files],
            download_link=output_file_url,
            message_class="success"
        )
    else:
        return render_template(
            'index.html',
            result=f"Matching percentage below threshold. Files not merged.",
            file_names=[os.path.basename(file.filename) for file in files],
            message_class="error"
        )

@app.route('/download/<filename>')
def download_file(filename):
    upload_folder = app.config['UPLOAD_FOLDER']
    file_path = os.path.join(upload_folder, filename)

    # Check if the file exists
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404

    return send_from_directory(upload_folder, filename)

if __name__ == '__main__':
    app.run(debug=True)

